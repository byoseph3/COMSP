#!/usr/bin/env python3
import argparse
import csv
import re
import json
from pathlib import Path
from urllib.parse import urlparse, parse_qs

import psycopg2
from psycopg2 import sql
from openpyxl import load_workbook

DEFAULT_ENV_PATHS = [Path('.') / '.env', Path('secrets') / '.env']

# Need to move this to a config file or something, but for now this is fine.
with open('secrets/reports_arr.json', 'r') as f:
    reports = json.load(f)['Reports']

def parse_env_file(path):
    values = {}
    with path.open('r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' not in line:
                continue
            key, value = line.split('=', 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            values[key] = value
    return values


def load_env(env_file=None):
    if env_file:
        path = Path(env_file)
        if not path.exists():
            raise FileNotFoundError(f"Environment file not found: {path}")
        return parse_env_file(path)

    for path in DEFAULT_ENV_PATHS:
        if path.exists():
            return parse_env_file(path)
    return {}


def parse_database_url(database_url):
    result = urlparse(database_url)
    user = result.username
    password = result.password
    host = result.hostname or 'localhost'
    port = result.port or 5432
    dbname = result.path.lstrip('/') if result.path else None
    return {
        'DB_USER': user,
        'DB_PASSWORD': password,
        'DB_HOST': host,
        'DB_PORT': str(port),
        'DB_NAME': dbname,
    }


def make_connection_params(env):
    if 'DATABASE_URL' in env and env.get('DB_USER') is None:
        env.update(parse_database_url(env['DATABASE_URL']))

    params = {
        'user': env.get('DB_USER'),
        'password': env.get('DB_PASSWORD'),
        'host': env.get('DB_HOST', 'localhost'),
        'port': env.get('DB_PORT', '5432'),
        'dbname': env.get('DB_NAME'),
    }

    missing = [k for k, v in params.items() if not v]
    if missing:
        raise ValueError(f"Missing database connection values: {', '.join(missing)}")

    return params


def normalize_column_name(value):
    cleaned = re.sub(r'[^a-zA-Z0-9_]', '_', value.strip().lower())
    cleaned = re.sub(r'_{2,}', '_', cleaned).strip('_')
    if not cleaned:
        raise ValueError(f"Invalid column name: {value!r}")
    return cleaned


def read_excel(path, sheet_name=None):
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def read_text_file(path, delimiter=None):
    with open(path, newline='', encoding='utf-8') as csvfile:
        sample = csvfile.read(2048)
        csvfile.seek(0)
        if delimiter is None:
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample, delimiters=',\t;|')
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ','
        reader = csv.reader(csvfile, delimiter=delimiter)
        rows = list(reader)
    if not rows:
        return [], []
    headers = [cell.strip() for cell in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def build_table_rows(headers, rows):
    columns = [normalize_column_name(header) for header in headers]
    unique = {}
    normalized = []
    for column in columns:
        count = unique.get(column, 0)
        unique[column] = count + 1
        normalized.append(f"{column}_{count + 1}" if count else column)

    data = []
    for row in rows:
        if len(row) < len(normalized):
            row = list(row) + [''] * (len(normalized) - len(row))
        data.append([None if cell == '' else str(cell) for cell in row[: len(normalized)]])
    return normalized, data


def create_or_replace_table(conn, schema, table_name, columns):
    with conn.cursor() as cur:
        cur.execute(
            sql.SQL('DROP TABLE IF EXISTS {schema}.{table}').format(
                schema=sql.Identifier(schema),
                table=sql.Identifier(table_name),
            )
        )
        col_defs = [sql.SQL('{} TEXT').format(sql.Identifier(col)) for col in columns]
        cur.execute(
            sql.SQL('CREATE TABLE {schema}.{table} ({cols})').format(
                schema=sql.Identifier(schema),
                table=sql.Identifier(table_name),
                cols=sql.SQL(', ').join(col_defs),
            )
        )
    conn.commit()

def initialize_attendance_columns(conn, schema, table_name):
    with conn.cursor() as cur:
        for report in reports:
            cur.execute(
                sql.SQL('ALTER TABLE {schema}.{table} ' \
                'ADD COLUMN IF NOT EXISTS {col} TEXT REFERENCES att_opt(att), ADD COLUMN {reason} TEXT').format(
                    schema=sql.Identifier(schema),
                    table=sql.Identifier(table_name),
                    col=sql.Identifier(f'{report} Alpha'),
                    reason=sql.Identifier(f'{report} Alpha Reason')
                )
            )
            cur.execute(
                sql.SQL('ALTER TABLE {schema}.{table} ' \
                'ADD COLUMN IF NOT EXISTS {col} TEXT REFERENCES att_opt(att), ADD COLUMN {reason} TEXT').format(
                    schema=sql.Identifier(schema),
                    table=sql.Identifier(table_name),
                    col=sql.Identifier(f'{report} Omega'),
                    reason=sql.Identifier(f'{report} Omega Reason')
                )
            )
    conn.commit()

def insert_rows(conn, schema, table_name, columns, rows):
    with conn.cursor() as cur:
        column_identifiers = [sql.Identifier(col) for col in columns]
        value_placeholders = sql.SQL(', ').join(sql.Placeholder() * len(columns))
        insert_query = sql.SQL('INSERT INTO {schema}.{table} ({cols}) VALUES ({values})').format(
            schema=sql.Identifier(schema),
            table=sql.Identifier(table_name),
            cols=sql.SQL(', ').join(column_identifiers),
            values=value_placeholders,
        )
        cur.executemany(insert_query, rows)
    conn.commit()


def parse_args():
    parser = argparse.ArgumentParser(description='Import users into a Postgres table from Excel or text files.')
    parser.add_argument('source', help='Path to an Excel (.xlsx) or text (.csv/.txt) file.')
    parser.add_argument('-t', '--table', default='users', help='Target table name in the database.')
    parser.add_argument('-s', '--schema', default='public', help='Target schema name. Default: public')
    parser.add_argument('-e', '--env-file', help='Path to the .env file with DB connection settings.')
    parser.add_argument('--sheet', help='Excel sheet name to import (defaults to the active sheet).')
    parser.add_argument('--delimiter', help='Delimiter for text files. If omitted, the script will attempt to detect it.')
    return parser.parse_args()


def main():
    args = parse_args()
    source = Path(args.source)
    if not source.exists():
        raise FileNotFoundError(f"Source file not found: {source}")

    env = load_env(args.env_file)
    params = make_connection_params(env)

    if source.suffix.lower() in {'.xlsx'}:
        headers, rows = read_excel(source, sheet_name=args.sheet)
    else:
        headers, rows = read_text_file(source, delimiter=args.delimiter)

    if not headers:
        raise ValueError('Input file has no column headers.')

    columns, table_rows = build_table_rows(headers, rows)
    print(f"Importing {len(table_rows)} row(s) into {args.schema}.{args.table}...")

    with psycopg2.connect(**params) as conn:
        create_or_replace_table(conn, args.schema, args.table, columns)
        if table_rows:
            insert_rows(conn, args.schema, args.table, columns, table_rows)

    # Final touchup, to alter table and include relations for attendance throughout the week.
    # TODO: Implement ALTER TABLE Statement
    initialize_attendance_columns(conn, args.schema, args.table)

    print(f"Finished. {len(table_rows)} row(s) imported into {args.schema}.{args.table}.")


if __name__ == '__main__':
    main()
